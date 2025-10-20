
# download PowerBi revenue detail report. Save it as "ab"
# Take the Mavenlink tracker and save the file as "ML"
# Final report will be called "Updated Revenue Report.xlsx"
# This code has been revised to also insert actual cost data for hardware. Replacing the 55% fixed rate with real data.
# Download the report "RAR Report" from PowerBi (remove all filters). Saved it as "Costs" and upload to Scripts file path.

import os
import pandas as pd

df_pb = pd.read_excel('ab.xlsx')

df_pb.columns = df_pb.columns.str.replace('[', '-')
df_pb.columns = df_pb.columns.str.replace(']', '-')
df_pb.columns = df_pb.columns.str.replace('\'','')

def xlookup(lookup_value, lookup_array, return_array, if_not_found:str = ''):
    match_value = return_array.loc[lookup_array == lookup_value]
    if match_value.empty:
        return f'"{lookup_value}" not found!' if if_not_found == '' else if_not_found

    else:
        return match_value.tolist()[0]





df_pb = df_pb[['Year', 'Month', 'JDE SO', 'Room Conversion Date', 'Account: Account Name', 'Account: Location',
            'Project Manager', 'Project', 'Project: Services Region', 'Revenue: Documentation Status', 'Revenue: Review Completed On',
              'JDE Account Name', 'JDE Account Number', 'Material Number', 'Order Line Category', 'Revenue Type', 'Sum of Net Price', 'Revenue (Gross)', 'Sum of Cost Required', 'HW/SW Only (Custom)' ]]

df_pb.to_excel('PowerBi.xlsx', sheet_name='PowerBi', index=False)


# Load the data from the "PowerBi.xlsx" and "ML.xlsx" files
df_powerbi = pd.read_excel('PowerBi.xlsx')
df_ml = pd.read_excel('ML.xlsx')

# Create a new Excel writer object for the combined file
with pd.ExcelWriter('Revenue Report.xlsx') as writer:
    # Write each DataFrame to a separate sheet
    df_powerbi.to_excel(writer, sheet_name='PowerBi', index=False)
    df_ml.to_excel(writer, sheet_name='ML', index=False)

print("Files have been combined into 'Revenue Report.xlsx'")


import os
import time

# Example to check file modification time
file_path = 'revenue_type.xlsx'
file_stats = os.stat(file_path)
print("Last modified time:", time.ctime(file_stats.st_mtime))


print(os.getcwd())

cwd = "C:\\Users\\bilonid.GLOBAL"

folder_name = 'finished_files'
file_name = "Final.xlsx"
full_path = os.path.join(cwd, folder_name, file_name)

# if not os.path.exists(os.path.join(cwd, folder_name, file_name)):
 #   os.makedirs(os.path.join(cwd, folder_name))

import openpyxl
from openpyxl.utils import get_column_letter

def create_variance_sheet_with_formulas(file_path):
    # Load the workbook and check for the required sheets
    workbook = openpyxl.load_workbook(file_path)
    powerbi_sheet = workbook["PowerBi"]
    ml_sheet = workbook["ML"]

    # Create a new sheet named "Variance" if it doesn't exist
    if "Variance" in workbook.sheetnames:
        variance_sheet = workbook["Variance"]
    else:
        variance_sheet = workbook.create_sheet("Variance")

    # Extract data from the first row onwards in PowerBi and ML sheets, ensuring all data is treated as strings
    powerbi_data = [str(cell.value).strip() for row in powerbi_sheet.iter_rows(min_row=1, min_col=3, max_col=3) for cell in row if cell.value is not None]
    ml_data = [str(cell.value).strip() for row in ml_sheet.iter_rows(min_row=1, min_col=4, max_col=4) for cell in row if cell.value is not None]

    # Combine data
    combined_data = powerbi_data + ml_data

    # Remove duplicates by converting to a set
    unique_data = set(combined_data)

    # Insert headers in the first row
    variance_sheet["A1"] = "SO#"
    variance_sheet["B1"] = "PowerBi"
    variance_sheet["C1"] = "Mavenlink"
    variance_sheet["D1"] = "Variance"

    # Write unique values to the "SO#" column in the Variance sheet starting from the second row
    for index, value in enumerate(sorted(unique_data), start=2):  # start at row 2 to keep the header
        cell_reference = f"A{index}"
        variance_sheet[cell_reference] = value

    # Add a "PowerBi" and "Mavenlink" columns with SUMIFS formulas starting from the second row
    for index in range(2, len(unique_data) + 2):  # adjust index for starting from row 2
        so_cell = f"A{index}"
        powerbi_formula = f'=SUMIFS(PowerBi!R:R, PowerBi!C:C, {so_cell})'
        mavenlink_formula = f'=SUMIFS(ML!H:H, ML!D:D, {so_cell})'
        variance_sheet[f"B{index}"] = powerbi_formula
        variance_sheet[f"C{index}"] = mavenlink_formula
        variance_formula = f"=B{index}-C{index}"
        variance_sheet[f"D{index}"] = variance_formula

    # Identify rows to delete where both PowerBi and Mavenlink columns show zero after formulas are evaluated
    max_row = variance_sheet.max_row
    rows_to_delete = []
    for row in range(2, max_row + 1):
        powerbi_cell = variance_sheet[f"B{row}"].value
        mavenlink_cell = variance_sheet[f"C{row}"].value
        if powerbi_cell == "0" and mavenlink_cell == "0":
            rows_to_delete.append(row)
    
    for row in sorted(rows_to_delete, reverse=True):
        variance_sheet.delete_rows(row)

    # Save the modified workbook
    workbook.save(file_path)

if __name__ == "__main__":
    # The user will need to specify the path to their Excel file here
    file_path = "Revenue Report.xlsx"
    create_variance_sheet_with_formulas(file_path)




import openpyxl

def perform_xlookup():
    # Load the workbook containing the "PowerBi" sheet
    main_workbook = openpyxl.load_workbook("Revenue Report.xlsx")
    powerbi_sheet = main_workbook["PowerBi"]
    
    # Load the workbook that contains the revenue type mappings
    lookup_workbook = openpyxl.load_workbook("revenue_type.xlsx")
    lookup_sheet = lookup_workbook.active  # Assuming there's only one sheet in this workbook

    # Build a dictionary for XLOOKUP
    # Key: Material Number, Value: Revenue Type
    revenue_type_lookup = {}
    for row in lookup_sheet.iter_rows(min_row=2, values_only=True):
        material_number = row[0]  # Assuming 'Material Number' is in the first column
        revenue_type = row[1]    # Assuming 'Revenue Type' is in the second column
        revenue_type_lookup[material_number] = revenue_type

    # Iterate over the "PowerBi" sheet and replace 'Revenue Type' based on 'Material Number'
    material_number_col_index = 14  # Example: assuming 'Material Number' is in column B
    revenue_type_col_index = 16     # Example: assuming 'Revenue Type' is in column C
    for row in powerbi_sheet.iter_rows(min_row=2):
        material_number = row[material_number_col_index-1].value
        if material_number in revenue_type_lookup:
            row[revenue_type_col_index-1].value = revenue_type_lookup[material_number]

    # Save the updated main workbook
    main_workbook.save("Updated Revenue Report.xlsx")

if __name__ == "__main__":
    perform_xlookup()


import openpyxl

def filter_canada_rows():
    # Load the workbook and access the 'PowerBi' sheet
    workbook = openpyxl.load_workbook("Updated Revenue Report.xlsx")
    powerbi_sheet = workbook["PowerBi"]
    
    # Create a new sheet for Canada rows if it does not exist
    if "PowerBi Canada" in workbook.sheetnames:
        canada_sheet = workbook["PowerBi Canada"]
        # If sheet exists, we clear it to ensure headers are copied correctly without duplication
        workbook.remove(canada_sheet)
        canada_sheet = workbook.create_sheet("PowerBi Canada")
    else:
        canada_sheet = workbook.create_sheet("PowerBi Canada")
    
    # Copy headers from 'PowerBi' to 'PowerBi Canada'
    headers = [cell.value for cell in powerbi_sheet[1]]
    canada_sheet.append(headers)

    # Determine the index for 'Project: Services Region'
    services_region_col_idx = headers.index("Project: Services Region") + 1

    # Initialize a list to keep track of rows to be removed
    rows_to_remove = []

    # Iterate through rows, find those with "Canada" in the specified column
    for row in powerbi_sheet.iter_rows(min_row=2):
        if "Canada" in str(row[services_region_col_idx - 1].value):
            rows_to_remove.append(row[0].row)
            canada_sheet.append([cell.value for cell in row])  # Append row to the Canada sheet

    # Remove rows containing "Canada" from the bottom up to avoid shifting indices
    for row_idx in reversed(rows_to_remove):
        powerbi_sheet.delete_rows(row_idx)

    # Save the modified workbook
    workbook.save("Updated Revenue Report.xlsx")

if __name__ == "__main__":
    filter_canada_rows()




import openpyxl
from openpyxl import Workbook

def create_excel_with_je_sheet(filename, lookup_filename):
    # Load the lookup workbook and prepare the lookup dictionaries
    lookup_workbook = openpyxl.load_workbook(lookup_filename)
    lookup_sheet = lookup_workbook.active
    description_lookup = {}
    type_lookup = {}

    # Populate the lookup dictionaries from the lookup sheet
    for row in lookup_sheet.iter_rows(min_row=1, values_only=True):
        account_number = row[0]
        type_value = row[1]
        description_value = row[2]
        description_lookup[account_number] = description_value
        type_lookup[account_number] = type_value

    # Try to open an existing workbook, otherwise create a new one
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()

    # Define headers
    headers = ["Account Number", "Amount", "Description", "Reference 2", "Type"]

    # Check if the 'JE' sheet exists and clear it
    if 'JE' in workbook.sheetnames:
        je_sheet = workbook['JE']
        je_sheet.delete_rows(1, je_sheet.max_row)  # Clear all existing content
    else:
        je_sheet = workbook.create_sheet('JE')

    # List of Account Numbers
    account_numbers = [
        "002840.4110.10", "002844.4110.10", "002842.4110.10",
        "002842.4110.10", "002843.4110.10", "002841.4110.10"
    ]

    # Populate the account numbers and perform xlookup for description and type
    for account in account_numbers:
        description = description_lookup.get(account, "")
        account_type = type_lookup.get(account, "")
        je_sheet.append([account, "", description, "", account_type])

    # Insert headers at the top of the sheet by shifting all data down
    je_sheet.insert_rows(1)
    for col, header in enumerate(headers, start=1):
        je_sheet.cell(row=1, column=col).value = header

    # Save the workbook
    workbook.save(filename)

# Function call (uncomment in actual use)
create_excel_with_je_sheet("Updated Revenue Report.xlsx", "lookup_gl.xlsx")





def update_cell(filename):
    workbook = openpyxl.load_workbook(filename)
    
    if 'JE' in workbook.sheetnames:
        je_sheet = workbook['JE']
        je_sheet['E4'] = "Professional Services"
        workbook.save(filename)
        
update_cell("Updated Revenue Report.xlsx")

import openpyxl

def perform_sumifs_in_excel(filename):
    workbook = openpyxl.load_workbook(filename)
   
    # Access both JE and PowerBi sheets
    je_sheet = workbook['JE']
    powerbi_sheet = workbook['PowerBi']
   
    # Define column indices for criteria and sum ranges
    criteria_col_index = 15  # Column P for criteria (16th column, 15 zero-based)
    sum_col_index = 17       # Column R for sum (18th column, 17 zero-based)

    # Create a dictionary to hold the sum values and the criteria from PowerBi
    revenue_data = {}
    for row in powerbi_sheet.iter_rows(min_row=2, values_only=True):
        revenue_type = row[criteria_col_index]
        # Convert revenue_gross to float to handle numerical data correctly
        try:
            revenue_gross = -float(row[sum_col_index])
        except ValueError:
            revenue_gross = 0  # Handle the case where conversion is not possible

        if revenue_type in revenue_data:
            revenue_data[revenue_type] += revenue_gross
        else:
            revenue_data[revenue_type] = revenue_gross

    # Update the JE sheet Amount column based on Type column using SUMIFS logic
    je_criteria_col_index = 4  # Column E (5th column, 4 zero-based) in JE for criteria
    amount_col_index = 1       # Column B (2nd column, 1 zero-based) in JE for amount
    for i, row in enumerate(je_sheet.iter_rows(min_row=2, max_row=je_sheet.max_row, min_col=je_criteria_col_index+1, max_col=je_criteria_col_index+1, values_only=True), start=2):
        je_criteria = row[0]
        sum_amount = revenue_data.get(je_criteria, 0)
        # Correct cell reference for row index
        je_sheet.cell(row=i, column=amount_col_index + 1).value = sum_amount

    # Save the updated workbook
    workbook.save(filename)

# Function call (uncomment in actual use)
perform_sumifs_in_excel("Updated Revenue Report.xlsx")





import openpyxl

def process_excel_file(file_path):
    # Load the workbook and sheets
    wb = openpyxl.load_workbook(file_path)
    sheet_je = wb["JE"]
    sheet_powerbi = wb["PowerBi"]
    
    # Extract unique values from "PowerBi" column C
    unique_values = set()
    for cell in sheet_powerbi['C'][1:]:  # Skip the header row
        unique_values.add(cell.value)

    # Paste unique values into "JE" column D starting at row 8
    d_column_start = 8
    for i, value in enumerate(unique_values, start=d_column_start):
        sheet_je[f'D{i}'].value = value

    # Find the first blank row in column B for inserting SUMIFS formula
    b_column_first_blank = d_column_start + len(unique_values)

    # Insert SUMIFS formulas in column B in the "JE" tab
    for index in range(d_column_start, b_column_first_blank):
        criteria_address = f'D{index}'
        formula = f'=SUMIFS(PowerBi!$R:$R, PowerBi!$C:$C, JE!{criteria_address})'
        sheet_je[f'B{index}'].value = formula

    # Paste "2.2460.40" in column A from row 8 until it reaches a blank row in column D
    for index in range(d_column_start, b_column_first_blank):
        sheet_je[f'A{index}'].value = "2.2460.40"

    # Insert "NC 100%" in column C starting from row 8 until the value in column A is not "2.2460.40"
    index = d_column_start
    while sheet_je[f'A{index}'].value == "2.2460.40":
        sheet_je[f'C{index}'].value = "NC 100%"
        index += 1
    
    # Save the modified workbook
    wb.save('Updated Revenue Report.xlsx')

# Uncomment the below line to execute the function with the specified file path
process_excel_file("Updated Revenue Report.xlsx")


import openpyxl

def process_excel_file(file_path):
    # Load the workbook and sheets
    wb = openpyxl.load_workbook(file_path)
    sheet_je = wb["JE"]

    # Check if "002840.5110.10" already exists in column A
    already_exists = False
    for cell in sheet_je['A']:
        if cell.value is not None and cell.value.strip() == "002840.5110.10":
            already_exists = True
            print("Found existing '002840.5110.10'. No new row added.")
            break

    if not already_exists:
        print("Did not find '002840.5110.10'. Adding new row.")
        # Find next blank row in "JE" sheet
        next_blank_row = 1
        while sheet_je[f'A{next_blank_row}'].value is not None:
            next_blank_row += 1

        # Paste "002840.5110.10" in column A and "ASC 606 RecCost Hardware" in column C
        sheet_je[f'A{next_blank_row}'].value = "002840.5110.10"
        sheet_je[f'C{next_blank_row}'].value = "ASC 606 RecCost Hardware"

        # Check if the second row of column E is "Hardware"
        if sheet_je['E2'].value == "Hardware":
            value_to_paste = -sheet_je['B2'].value * 0.55
        else:
            value_to_paste = "NA"

        sheet_je[f'B{next_blank_row}'].value = value_to_paste

        next_blank_row += 1
        sheet_je[f'A{next_blank_row}'].value = "2.1399.40"
        previous_b = f'B{next_blank_row - 1}'
        sheet_je[f'B{next_blank_row}'].value = f"=-{previous_b}"
        sheet_je[f'C{next_blank_row}'].value = "Rec. Cost from RAR"

    # Save the modified workbook
    wb.save(file_path)

# Uncomment the below line to execute the function with the specified file path
process_excel_file("Updated Revenue Report.xlsx")





import openpyxl

def create_je_canada_sheet(main_file_path, lookup_file_path):
    # Load the main workbook and lookup workbook
    wb = openpyxl.load_workbook(main_file_path)
    lookup_wb = openpyxl.load_workbook(lookup_file_path)
    lookup_sheet = lookup_wb["US"]  # Adjusted for the "US" sheet
    
    # Create new sheet 'JE Canada' or clear it if it already exists
    if "JE Canada" in wb.sheetnames:
        je_canada = wb["JE Canada"]
        je_canada.delete_rows(1, je_canada.max_row)
    else:
        je_canada = wb.create_sheet("JE Canada")
    
    # Set headers
    headers = ["Account Number", "Amount", "Description", "Reference 2", "Type"]
    je_canada.append(headers)
    
    # Populate the Account Number column with predefined account numbers
    account_numbers = ["003840.4110.30", "003844.4110.30", "003842.4110.30", "003842.4110.30", "003843.4110.30", "003841.4110.30"]
    for account in account_numbers:
        je_canada.append([account])

    # Build lookup dictionary from the lookup sheet
    lookup_dict_type = {}
    lookup_dict_desc = {}
    for row in lookup_sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers are in the first row
        lookup_dict_type[row[0]] = row[1]
        lookup_dict_desc[row[0]] = row[2]

    # Use the dictionary to assign values to 'Type' and 'Description' in 'JE Canada'
    for idx, account in enumerate(account_numbers, start=2):
        je_canada.cell(row=idx, column=5).value = lookup_dict_type.get(account, "NA")  # Column E for 'Type'
        je_canada.cell(row=idx, column=3).value = lookup_dict_desc.get(account, "NA")  # Column C for 'Description'
    
    # Save the workbook
    wb.save(main_file_path)

# Usage example
create_je_canada_sheet("Updated Revenue Report.xlsx", "lookup_gl.xlsx")



def update_cell(filename):
    workbook = openpyxl.load_workbook(filename)
    
    if 'JE Canada' in workbook.sheetnames:
        je_sheet = workbook['JE Canada']
        je_sheet['E5'] = "Room Package"
        workbook.save(filename)
        
update_cell("Updated Revenue Report.xlsx")

import openpyxl

def add_sumifs_to_je_canada(workbook_path):
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Ensure "JE Canada" sheet exists or create it
    if "JE Canada" not in wb.sheetnames:
        je_canada = wb.create_sheet("JE Canada")
        # Assuming we need to set headers or initialize something else if the sheet is newly created
        je_canada.append(["Account Number", "Amount", "Description", "Reference 2", "Type"])
    else:
        je_canada = wb["JE Canada"]

    # Ensure "PowerBi Canada" sheet exists
    if "PowerBi Canada" not in wb.sheetnames:
        print("Error: 'PowerBi Canada' sheet does not exist in the workbook.")
        return

    # Assuming account numbers or related data are already filled in JE Canada and we start formulas from row 2
    # Check and add formulas only if there's data in column E to use as criteria
    max_row = je_canada.max_row
    for row in range(2, max_row + 1):
        criteria_cell = je_canada.cell(row=row, column=5)  # Column E for criteria
        if criteria_cell.value is not None:
            sumifs_formula = f"=-SUMIFS('PowerBi Canada'!$R:$R, 'PowerBi Canada'!$P:$P, {criteria_cell.coordinate})"
            je_canada.cell(row=row, column=2).value = sumifs_formula  # Column B for the SUMIFS formula

    # Save changes to the workbook
    wb.save(workbook_path)

# Usage example
add_sumifs_to_je_canada("Updated Revenue Report.xlsx")


import openpyxl

def update_je_canada_with_data(workbook_path):
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Access or create the necessary sheets
    if "JE Canada" not in wb.sheetnames:
        je_canada = wb.create_sheet("JE Canada")
    else:
        je_canada = wb["JE Canada"]

    if "PowerBi Canada" not in wb.sheetnames:
        print("Error: 'PowerBi Canada' sheet does not exist in the workbook.")
        return
    else:
        powerbi_canada = wb["PowerBi Canada"]

    # Extract unique values from column C of 'PowerBi Canada'
    unique_values = set()
    for cell in powerbi_canada['C']:
        if cell.value is not None and cell.row != 1:  # Exclude header and None values
            unique_values.add(cell.value)

    # Paste unique values into column D of 'JE Canada' starting from row 8
    start_row = 8
    for i, value in enumerate(unique_values, start=start_row):
        je_canada.cell(row=i, column=4).value = value  # Column 4 is column D

    # Add SUMIFS formula in column B of 'JE Canada'
    # Assuming formulas should be applied from row 7 onwards
    for i in range(start_row, start_row + len(unique_values)):
        criteria_cell = je_canada.cell(row=i, column=4)  # Criteria in column D
        sumifs_formula = f"=SUMIFS('PowerBi Canada'!$R:$R, 'PowerBi Canada'!$C:$C, {criteria_cell.coordinate})"
        je_canada.cell(row=i, column=2).value = sumifs_formula  # Column B for the SUMIFS formula

    # Save changes to the workbook
    wb.save(workbook_path)

# Usage example
update_je_canada_with_data("Updated Revenue Report.xlsx")


import openpyxl

def update_sheet_based_on_condition(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
   
    # Check if the 'JE Canada' sheet exists
    if 'JE Canada' in workbook.sheetnames:
        sheet = workbook['JE Canada']
    else:
        print("Sheet 'JE Canada' does not exist.")
        return
   
    # Iterate through each row in the sheet starting from the first row with potential data
    for row in range(2, sheet.max_row + 1):
        # Check if there's a value in column D (column index 4)
        if sheet.cell(row=row, column=4).value is not None:
            # Insert "3.2460.40" in column A and "NC 100% Revenue" in column C
            sheet.cell(row=row, column=1).value = "3.2460.40"
            sheet.cell(row=row, column=3).value = "NC 100% Revenue"
   
    # Save the workbook with changes
    workbook.save(filename)
    print("Workbook updated successfully based on conditions in column D.")

# Function call (uncomment in actual use)
update_sheet_based_on_condition("Updated Revenue Report.xlsx")


import openpyxl

def update_sheet_with_data(filename):
    # Load the workbook, assuming it exists.
    workbook = openpyxl.load_workbook(filename)

    # Access the 'JE Canada' sheet
    if 'JE Canada' in workbook.sheetnames:
        sheet = workbook['JE Canada']
    else:
        print("Sheet 'JE Canada' does not exist. Creating a new sheet.")
        sheet = workbook.create_sheet('JE Canada')

    # Find the next blank row in column A
    row_index = 1
    while sheet[f'A{row_index}'].value is not None:
        row_index += 1

    # Set the values in the next blank row
    sheet[f'A{row_index}'].value = "003840.5110.30"
    sheet[f'C{row_index}'].value = "ASC606 Rec Cost Hardware"
    # Insert the formula that calculates -B2 * 0.55 in column B
    sheet[f'B{row_index}'].value = "=-B2 * 0.55"

    # Save the workbook
    workbook.save(filename)
    print(f"Data added successfully in row {row_index} of 'JE Canada'.")

# Function call (uncomment in actual use)
update_sheet_with_data("Updated Revenue Report.xlsx")



import openpyxl

def update_je_canada(filename):
    # Load the workbook, assuming it exists.
    workbook = openpyxl.load_workbook(filename)

    # Access or create the 'JE Canada' sheet
    if 'JE Canada' in workbook.sheetnames:
        sheet = workbook['JE Canada']
    else:
        print("Sheet 'JE Canada' does not exist. Creating a new sheet.")
        sheet = workbook.create_sheet('JE Canada')

    # Find the next blank row in column A
    row_index = 1
    while sheet[f'A{row_index}'].value is not None:
        row_index += 1

    # Paste the value "3.1399.40" in column A
    sheet[f'A{row_index}'].value = "3.1399.40"

    # Insert a formula in column B that equals the negative of the cell immediately above it
    if row_index > 1:  # Ensure there is a row above to reference
        sheet[f'B{row_index}'].value = f"=-B{row_index - 1}"
    else:
        sheet[f'B{row_index}'].value = "0"  # If it's the first row, no above cell to reference

    # Paste "Rec. Cost from RAR" in column C
    sheet[f'C{row_index}'].value = "Rec. Cost from RAR"

    # Save the workbook
    workbook.save(filename)
    print(f"Data added successfully in row {row_index} of 'JE Canada'.")

# Function call (uncomment in actual use)
update_je_canada("Updated Revenue Report.xlsx")


# New Addition

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated Revenue Report.xlsx')

# Get the relevant sheets
variance_sheet = workbook['Variance']
powerbi_sheet = workbook['PowerBi Canada']

# Create a new sheet for "Variance Canada" if it doesn't already exist
if 'Variance Canada' not in workbook.sheetnames:
    variance_canada_sheet = workbook.create_sheet('Variance Canada')
else:
    variance_canada_sheet = workbook['Variance Canada']

# Copy headers from Variance sheet to Variance Canada sheet
for col_num, cell in enumerate(variance_sheet[1], 1):
    variance_canada_sheet.cell(row=1, column=col_num, value=cell.value)

# Identify the column indices for "SO#" in Variance and "JDE SO-JDE SO-" in PowerBi Canada
variance_headers = [cell.value for cell in variance_sheet[1]]
powerbi_headers = [cell.value for cell in powerbi_sheet[1]]

so_col_index = variance_headers.index("SO#") + 1
jde_so_col_index = powerbi_headers.index("JDE SO") + 1

# Function to find matching rows and perform copying and deletion
def match_and_transfer_rows():
    variance_rows_to_delete = []
    for var_row in variance_sheet.iter_rows(min_row=2, values_only=True):
        so_value = str(var_row[so_col_index - 1]).strip()
        for pbi_row in powerbi_sheet.iter_rows(min_row=2, values_only=True):
            if so_value == str(pbi_row[jde_so_col_index - 1]).strip():
                variance_canada_sheet.append(var_row)
                variance_rows_to_delete.append(var_row)
                break

    # Delete matching rows from Variance sheet
    for row in variance_rows_to_delete:
        for cell in variance_sheet.iter_rows(min_row=2):
            if str(cell[so_col_index - 1].value).strip() == str(row[so_col_index - 1]).strip():
                variance_sheet.delete_rows(cell[0].row)
                break

# Call the function to match and transfer rows
match_and_transfer_rows()

# Save the updated workbook
workbook.save('Updated Revenue Report.xlsx')



# New Addition to add Canada variance tab.
import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated Revenue Report.xlsx')

# Get the relevant sheets
variance_canada_sheet = workbook['Variance Canada']
powerbi_sheet = workbook['PowerBi Canada']
ml_sheet = workbook['ML']

# Find the column indices
variance_canada_headers = [cell.value for cell in variance_canada_sheet[1]]
so_col_index = variance_canada_headers.index("SO#") + 1

# Define column headers and insert them if they don't exist
if 'PowerBi' not in variance_canada_headers:
    variance_canada_headers.append('PowerBi')
    variance_canada_sheet.cell(row=1, column=len(variance_canada_headers), value='PowerBi')

if 'Mavenlink' not in variance_canada_headers:
    variance_canada_headers.append('Mavenlink')
    variance_canada_sheet.cell(row=1, column=len(variance_canada_headers), value='Mavenlink')

if 'Variance' not in variance_canada_headers:
    variance_canada_headers.append('Variance')
    variance_canada_sheet.cell(row=1, column=len(variance_canada_headers), value='Variance')

# Get column indices after adding new headers
powerbi_col_index = variance_canada_headers.index('PowerBi') + 1
mavenlink_col_index = variance_canada_headers.index('Mavenlink') + 1
variance_col_index = variance_canada_headers.index('Variance') + 1

# Iterate over the rows in Variance Canada sheet starting from the second row
for row in range(2, variance_canada_sheet.max_row + 1):
    so_value = variance_canada_sheet.cell(row=row, column=so_col_index).value

    # SUMIFS formula for the PowerBi column
    sumifs_formula = f'=SUMIFS(\'PowerBi Canada\'!R:R, \'PowerBi Canada\'!C:C, A{row})'
    variance_canada_sheet.cell(row=row, column=powerbi_col_index, value=sumifs_formula)

    # SUMIFS formula for the Mavenlink column
    xlookup_formula = f'=SUMIFS(\'ML\'!G:G, \'ML\'!D:D, A{row})'
    variance_canada_sheet.cell(row=row, column=mavenlink_col_index, value=xlookup_formula)

    # Variance formula
    variance_formula = f'={openpyxl.utils.get_column_letter(mavenlink_col_index)}{row} - {openpyxl.utils.get_column_letter(powerbi_col_index)}{row}'
    variance_canada_sheet.cell(row=row, column=variance_col_index, value=variance_formula)

# Save the updated workbook
workbook.save('Updated Revenue Report.xlsx')


# Additional layer to ensure formulas are calculating correctly.

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated Revenue Report.xlsx')

# Get the relevant sheets
variance_sheet = workbook['Variance']
powerbi_sheet = workbook['PowerBi']
ml_sheet = workbook['ML']

# Identify the column indices for "PowerBi" and "Mavenlink" in Variance sheet
variance_headers = [cell.value for cell in variance_sheet[1]]

if 'PowerBi' not in variance_headers:
    variance_headers.append('PowerBi')
    variance_sheet.cell(row=1, column=len(variance_headers), value='PowerBi')

if 'Mavenlink' not in variance_headers:
    variance_headers.append('Mavenlink')
    variance_sheet.cell(row=1, column=len(variance_headers), value='Mavenlink')

# Get column indices after adding new headers
powerbi_col_index = variance_headers.index('PowerBi') + 1
mavenlink_col_index = variance_headers.index('Mavenlink') + 1
criteria_col_index = variance_headers.index("SO#") + 1

# Iterate over the rows in Variance sheet starting from the second row
for row in range(2, variance_sheet.max_row + 1):
    # SUMIFS formula for the PowerBi column
    sumifs_powerbi_formula = f'=SUMIFS(\'PowerBi\'!R:R, \'PowerBi\'!C:C, A{row})'
    variance_sheet.cell(row=row, column=powerbi_col_index, value=sumifs_powerbi_formula)

    # SUMIFS formula for the Mavenlink column
    sumifs_mavenlink_formula = f'=SUMIFS(\'ML\'!G:G, \'ML\'!D:D, A{row})'
    variance_sheet.cell(row=row, column=mavenlink_col_index, value=sumifs_mavenlink_formula)

# Save the updated workbook
workbook.save('Updated Revenue Report.xlsx')


# Additional layer to ensure variance column is calculating correctly.

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated Revenue Report.xlsx')

# Get the relevant sheet
variance_sheet = workbook['Variance']

# Identify the column index for "Variance" and ensure it exists
variance_headers = [cell.value for cell in variance_sheet[1]]

if 'Variance' not in variance_headers:
    variance_headers.append('Variance')
    variance_sheet.cell(row=1, column=len(variance_headers), value='Variance')

# Get column indices after adding the new header
variance_col_index = variance_headers.index('Variance') + 1

# Iterate over the rows in Variance sheet starting from the second row
for row in range(2, variance_sheet.max_row + 1):
    # Formula to subtract the value in column B from the value in column C
    formula = f'=C{row} - B{row}'
    variance_sheet.cell(row=row, column=variance_col_index, value=formula)

# Save the updated workbook
workbook.save('Updated Revenue Report.xlsx')


import pandas as pd

# Load the Excel file
file_path = 'Costs.xlsx'
df = pd.read_excel(file_path)

# Specify the columns to keep
columns_to_keep = [
    "SO", "CUSTOMER_NUMBER", "CUSTOMER_NAME", "COUNTRY_CODE",
    "INVOICE_DATE", "JDE_SRP1_DESCRIPTION", "MATERIAL_NUMBER", "Sum of COST_CONSUMED"
]

# Filter the DataFrame to keep only the specified columns
df_clean = df[columns_to_keep]

# Save the cleaned DataFrame to a new Excel file
output_file_path = 'Costs_Clean.xlsx'
df_clean.to_excel(output_file_path, index=False)

print(f"File saved as {output_file_path}")


from openpyxl import load_workbook

# Load the source workbook and the target workbook
source_workbook = load_workbook('Updated Revenue Report.xlsx')
target_workbook = load_workbook('Costs_Clean.xlsx')

# Get the sheet to be copied from the source workbook
source_sheet = source_workbook['PowerBi']

# Create a new sheet in the target workbook with the same name
target_sheet = target_workbook.create_sheet(title='PowerBi')

# Copy the contents from the source sheet to the target sheet
for row in source_sheet.iter_rows(values_only=True):
    target_sheet.append(row)

# Save the target workbook
target_workbook.save('Costs_Clean.xlsx')



from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the workbook
workbook = load_workbook('Costs_Clean.xlsx')

# Rename "Sheet1" to "All Costs"
workbook['Sheet1'].title = 'All_Costs'


# Access the necessary sheets
allcosts = workbook['All_Costs']
powerbi_sheet = workbook['PowerBi']

# Create a new sheet called "Totals"
totals_sheet = workbook.create_sheet(title='COS')

# Set headers in the "Totals" sheet
totals_sheet['A1'] = 'SO'
totals_sheet['B1'] = 'Sum of Cost'

# Get unique values from column C of the "PowerBi" sheet
powerbi_column_c = list(set(cell.value for cell in powerbi_sheet['C'] if cell.value is not None))

# Populate the "SO" column with unique values from "PowerBi" column C
for index, value in enumerate(powerbi_column_c, start=2):
    totals_sheet[f'A{index}'] = value
    # Create SUMIFS formula
    totals_sheet[f'B{index}'] = f'=SUMIFS(All_Costs!H:H, All_Costs!A:A, COS!A{index})'

# Save the modified workbook as "Final Costs"
workbook.save('Final Costs.xlsx')



from openpyxl import load_workbook

# Load the "Final Costs" workbook
final_costs_workbook = load_workbook('Final Costs.xlsx')

# Save the changes in the "Final Costs" workbook
final_costs_workbook.save('Final Costs.xlsx')

# Re-load the updated "Final Costs" workbook to ensure changes are saved
final_costs_workbook = load_workbook('Final Costs.xlsx')

# Access the sheets
all_costs_sheet = final_costs_workbook['All_Costs']
totals_sheet = final_costs_workbook['COS']

# Load the "Updated Revenue Report" workbook
updated_revenue_workbook = load_workbook('Updated Revenue Report.xlsx')

# Copy the "All Costs" sheet
new_all_costs_sheet = updated_revenue_workbook.create_sheet(title='All_Costs')
for row in all_costs_sheet.iter_rows(values_only=True):
    new_all_costs_sheet.append(row)

# Copy the "Totals" sheet
new_totals_sheet = updated_revenue_workbook.create_sheet(title='COS')
for row in totals_sheet.iter_rows(values_only=True):
    new_totals_sheet.append(row)

# Save the updated workbook
updated_revenue_workbook.save('Updated Revenue Report.xlsx')


from openpyxl import load_workbook

# Load the "Updated Revenue Report" workbook
workbook = load_workbook('Updated Revenue Report.xlsx')

# Access the "COS" sheet
cos_sheet = workbook['COS']

# Copy and hardcode the values from column B to column C
for row in cos_sheet.iter_rows(min_row=1, min_col=2, max_col=2):
    for cell in row:
        cos_sheet[f'C{cell.row}'].value = cell.value

# Set the formula for the sum in cell D1 of the "COS" sheet
cos_sheet['D1'].value = '=SUM(C:C)'

# Save the updated workbook
workbook.save('Updated Revenue Report.xlsx')


from openpyxl import load_workbook

# Load the "Updated Revenue Report" workbook
workbook = load_workbook('Updated Revenue Report.xlsx')

# Access the necessary sheets
je_sheet = workbook['JE']
cos_sheet = workbook['COS']

# Define the formula to be inserted
formula = '=COS!D1'

# Find the cell in column B of the "JE" sheet where column A is "002840.5110.10"
for row in je_sheet.iter_rows(min_row=1, max_col=2):
    if row[0].value == "002840.5110.10":
        # Replace the value of the cell in column B with the formula
        row[1].value = formula
        # Replace the value of the cell immediately below it with the negative formula
        next_row_index = row[1].row + 1
        je_sheet[f'B{next_row_index}'].value = f'=-COS!D1'
        break

# Save the updated workbook
workbook.save('Updated Revenue Report.xlsx')




import openpyxl

def insert_sumifs_formula(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    
    # Access the JE sheet
    je_sheet = wb['JE']
    
    # Access the PowerBi sheet
    # powerbi_sheet = wb['PowerBi']  # This line is optional as we only reference this sheet in the formula

    # Define the range where formulas will be inserted
    formula_range = je_sheet['B2':'B7']
    
    # Insert the SUMIFS formula
    for row in formula_range:
        for cell in row:
            # Formula referencing the PowerBi sheet and matching on column E of JE
            cell.value = f"=-SUMIFS(PowerBi!R:R, PowerBi!P:P, JE!E{cell.row})"
    
    # Save the workbook
    wb.save(filename)
    print(f"Formulas inserted and workbook saved as {filename}")

# Usage example
insert_sumifs_formula('Updated Revenue Report.xlsx')





