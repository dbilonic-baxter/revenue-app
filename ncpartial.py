# download PowerBi revenue detail report. Save it as "powerbipartial.xlsx"
# Take the Mavenlink tracker and save the file as "MLPARTIAL.xlsx"
# Update the "revenue_type" file if certain material #'s are not mapping
# Final file will be "Updated_Partial_Revenue.xlsx"
# Check why multiplication is off in some rows. Check to delete zero rows.

import os
import pandas as pd

df_pb = pd.read_excel('powerbipartial.xlsx')

import os
import time

# Example to check file modification time
file_path = 'powerbipartial.xlsx'
file_stats = os.stat(file_path)
print("Last modified time:", time.ctime(file_stats.st_mtime))

df_pb.columns = df_pb.columns.str.replace('[', '-')
df_pb.columns = df_pb.columns.str.replace(']', '-')
df_pb.columns = df_pb.columns.str.replace('\'','')


parts_df = pd.read_excel('Parts.xlsx')

def xlookup(lookup_value, lookup_array, return_array, if_not_found:str = ''):
    match_value = return_array.loc[lookup_array == lookup_value]
    if match_value.empty:
        return f'"{lookup_value}" not found!' if if_not_found == '' else if_not_found

    else:
        return match_value.tolist()[0]

df_pb['Part Name'] = df_pb['Material Number'].apply(xlookup, args = (parts_df['Item Number'], parts_df['Description']))

df_pb.columns

df_pb = df_pb[['Year', 'Month', 'JDE SO', 'Room Conversion Date', 'Account: Account Name', 'Account: Location',
            'Project Manager', 'Project', 'Project: Services Region', 'Revenue: Documentation Status', 'Revenue: Review Completed On',
              'JDE Account Name', 'JDE Account Number', 'Material Number', 'Order Line Category', 'Revenue Type', 'Sum of Net Price', 'Revenue (Gross)', 'Part Name' ]]

df_pb['S/B Revenue'] = 0
df_pb['ML Percentage Recognized'] = 0.0


df_pb.to_excel('powerbipartial.xlsx', index=False)

import pandas as pd

def combine_excel_sheets(source_files, output_file):
    # Create a new Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
    
    # Process each source file
    for file_name in source_files:
        # Load all sheets from the current file
        xls = pd.ExcelFile(file_name)
        for sheet_name in xls.sheet_names:
            # Read each sheet
            df = xls.parse(sheet_name)
            # Define new sheet name based on the file and sheet name
            if file_name == "powerbipartial.xlsx" and sheet_name == "Sheet1":
                new_sheet_name = "PowerBi"
            elif file_name == "MLPARTIAL.xlsx" and sheet_name == "Sheet1":
                new_sheet_name = "Mavenlink"
            else:
                new_sheet_name = f"{file_name.split('.')[0]}_{sheet_name}"
            # Write each sheet to the new file with the determined sheet name
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)
    
    # After copying all sheets, add the Variance sheet
    variance_df = pd.DataFrame(columns=["SO#", "PowerBi", "Mavenlink", "Variance"])
    variance_df.to_excel(writer, sheet_name='Variance', index=False)
    
    # Close the writer to save the workbook
    writer.close()
    
    print(f"All sheets combined and saved to {output_file}")

# List of source Excel files
source_files = ["powerbipartial.xlsx", "MLPARTIAL.xlsx"]
output_file = "partialrevenue.xlsx"



# Load ML (case/space tolerant)
df_ml_raw = pd.read_excel("MLPARTIAL.xlsx", sheet_name="Summary", header=2)
df_ml_raw.columns = df_ml_raw.columns.str.strip()

# find the "Special Action" column, regardless of exact casing/spaces
special_action_col = next(
    (c for c in df_ml_raw.columns if c.replace(" ", "").lower() == "specialaction"),
    None
)
if special_action_col is None:
    raise ValueError("Could not find 'Special Action' column in MLPARTIAL.xlsx -> Summary")

# filter down to "Partial Rec"
df_ml_raw[special_action_col] = df_ml_raw[special_action_col].astype(str).str.strip()
df_ml_filt = df_ml_raw[df_ml_raw[special_action_col].str.casefold() == "partial rec"]

# === Keep ONLY these columns, in this EXACT order ===
# NOTE: column 4 (D) must be Sales Order Number, and column 6 (F) must be the % used later.
#       This preserves your later logic where you read D and F from the Mavenlink sheet.
KEEP_ORDER = [
    "Project Manager",           # A
    "Account",                   # B
    "Install ID (Task)",         # C
    "Sales Order Number",        # D  <-- used later as the lookup key
    special_action_col,          # E
    "Total % Completed",         # F  <-- used later as the value
    "Total SO Extended $ Amount",# G
    "Total Rec'd"                # H
]

# build a mapping that tolerates case/space differences
norm = {c: c for c in df_ml_filt.columns}
def find(colname):
    target = colname.replace(" ", "").lower()
    for c in df_ml_filt.columns:
        if c.replace(" ", "").lower() == target:
            return c
    return None

resolved_cols = []
for name in KEEP_ORDER:
    if name == special_action_col:
        resolved_cols.append(special_action_col)
    else:
        col_found = find(name)
        if col_found is None:
            raise ValueError(f"Required column not found in ML: '{name}'")
        resolved_cols.append(col_found)

df_ml_final = df_ml_filt[resolved_cols].copy()

# Rename the dynamic special action back to a stable name (optional)
df_ml_final.rename(columns={special_action_col: "Special Action"}, inplace=True)

# Overwrite MLPARTIAL.xlsx with ONLY the kept columns (sheet name defaults to Sheet1)
df_ml_final.to_excel("MLPARTIAL.xlsx", index=False)



















# Call the function
combine_excel_sheets(source_files, output_file)



import openpyxl

def perform_lookup_and_calculate(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Ensure both "PowerBi" and "Mavenlink" sheets exist
    if "PowerBi" not in workbook.sheetnames or "Mavenlink" not in workbook.sheetnames:
        print("One or both required sheets are missing.")
        return

    # Access the sheets
    powerbi_sheet = workbook["PowerBi"]
    mavenlink_sheet = workbook["Mavenlink"]

    # Create a dictionary from Mavenlink for the lookup
    # Key: Column D values, Value: Column F values
    mavenlink_dict = {}
    for row in range(2, mavenlink_sheet.max_row + 1):
        key = mavenlink_sheet.cell(row, 4).value  # Column D
        value = mavenlink_sheet.cell(row, 6).value  # Column F
        mavenlink_dict[key] = value

    # Use the dictionary to perform lookup and populate column U in PowerBi
    for row in range(2, powerbi_sheet.max_row + 1):
        lookup_value = powerbi_sheet.cell(row, 3).value  # Column C
        result_value = mavenlink_dict.get(lookup_value, 0)
        powerbi_sheet.cell(row, 21, result_value)  # Column U

        # Insert multiplication formula in column T
        powerbi_sheet.cell(row, 20).value = f"=U{row} * R{row}"  # Column T

    # Save the workbook
    workbook.save(file_path)
    print(f"Workbook saved with updated calculations at {file_path}")

# File path to the workbook
file_path = "partialrevenue.xlsx"

# Function call
perform_lookup_and_calculate(file_path)




import openpyxl
from openpyxl.utils import get_column_letter

def create_variance_sheet_with_formulas(file_path):
    # Load the workbook and check for the required sheets
    workbook = openpyxl.load_workbook(file_path)
    powerbi_sheet = workbook["PowerBi"]
    ml_sheet = workbook["Mavenlink"]

    # Create a new sheet named "Variance" if it doesn't exist
    if "Variance" in workbook.sheetnames:
        variance_sheet = workbook["Variance"]
    else:
        variance_sheet = workbook.create_sheet("Variance")

    # Extract data from the first row onwards in PowerBi and ML sheets, ensuring all data is treated as strings
    powerbi_data = [str(cell.value).strip() for row in powerbi_sheet.iter_rows(min_row=1, min_col=3, max_col=3) for cell in row if cell.value is not None]
    ml_data = [str(cell.value).strip() for row in ml_sheet.iter_rows(min_row=1, min_col=4, max_col=4) for cell in row if cell.value is not None]

    # Combine data
    combined_data = powerbi_data + ml_data

    # Remove duplicates by converting to a set
    unique_data = set(combined_data)

    # Insert headers in the first row
    # variance_sheet["A1"] = "SO#"
    # variance_sheet["B1"] = "PowerBi"
    # variance_sheet["C1"] = "Mavenlink"
    # variance_sheet["D1"] = "Variance"

    # Write unique values to the "SO#" column in the Variance sheet starting from the second row
    for index, value in enumerate(sorted(unique_data), start=2):  # start at row 2 to keep the header
        cell_reference = f"A{index}"
        variance_sheet[cell_reference] = value

    # Add a "PowerBi" and "Mavenlink" columns with SUMIFS formulas starting from the second row
    for index in range(2, len(unique_data) + 2):  # adjust index for starting from row 2
        so_cell = f"A{index}"
        powerbi_formula = f'=SUMIFS(PowerBi!T:T, PowerBi!C:C, {so_cell})'
        mavenlink_formula = f'=SUMIFS(Mavenlink!H:H, Mavenlink!D:D, {so_cell})'
        variance_sheet[f"B{index}"] = powerbi_formula
        variance_sheet[f"C{index}"] = mavenlink_formula
        variance_formula = f"=B{index}-C{index}"
        variance_sheet[f"D{index}"] = variance_formula

    # Identify rows to delete where both PowerBi and Mavenlink columns show zero after formulas are evaluated
    max_row = variance_sheet.max_row
    rows_to_delete = []
    for row in range(2, max_row + 1):
        powerbi_cell = variance_sheet[f"B{row}"].value
        mavenlink_cell = variance_sheet[f"C{row}"].value
        if powerbi_cell == "0" and mavenlink_cell == "0":
            rows_to_delete.append(row)
    
    for row in sorted(rows_to_delete, reverse=True):
        variance_sheet.delete_rows(row)

    # Save the modified workbook
    workbook.save(file_path)

if __name__ == "__main__":
    # The user will need to specify the path to their Excel file here
    file_path = "partialrevenue.xlsx"
    create_variance_sheet_with_formulas(file_path)






import openpyxl

def perform_xlookup():
    # Load the workbook containing the "PowerBi" sheet
    main_workbook = openpyxl.load_workbook("partialrevenue.xlsx")
    powerbi_sheet = main_workbook["PowerBi"]
    
    # Load the workbook that contains the revenue type mappings
    lookup_workbook = openpyxl.load_workbook("revenue_type.xlsx")
    lookup_sheet = lookup_workbook.active  # Assuming there's only one sheet in this workbook

    # Build a dictionary for XLOOKUP
    # Key: Material Number, Value: Revenue Type
    revenue_type_lookup = {}
    for row in lookup_sheet.iter_rows(min_row=2, values_only=True):
        material_number = row[0]  # Assuming 'Material Number' is in the first column
        revenue_type = row[1]    # Assuming 'Revenue Type' is in the second column
        revenue_type_lookup[material_number] = revenue_type

    # Iterate over the "PowerBi" sheet and replace 'Revenue Type' based on 'Material Number'
    material_number_col_index = 14  # Example: assuming 'Material Number' is in column B
    revenue_type_col_index = 16     # Example: assuming 'Revenue Type' is in column C
    for row in powerbi_sheet.iter_rows(min_row=2):
        material_number = row[material_number_col_index-1].value
        if material_number in revenue_type_lookup:
            row[revenue_type_col_index-1].value = revenue_type_lookup[material_number]

    # Save the updated main workbook
    main_workbook.save("Updated_Partial_Revenue.xlsx")

if __name__ == "__main__":
    perform_xlookup()


import openpyxl

def filter_canada_rows():
    # Load the workbook and access the 'PowerBi' sheet
    workbook = openpyxl.load_workbook("Updated_Partial_Revenue.xlsx")
    powerbi_sheet = workbook["PowerBi"]
    
    # Create a new sheet for Canada rows if it does not exist
    if "PowerBi Canada" in workbook.sheetnames:
        canada_sheet = workbook["PowerBi Canada"]
        # If sheet exists, we clear it to ensure headers are copied correctly without duplication
        workbook.remove(canada_sheet)
        canada_sheet = workbook.create_sheet("PowerBi Canada")
    else:
        canada_sheet = workbook.create_sheet("PowerBi Canada")
    
    # Copy headers from 'PowerBi' to 'PowerBi Canada'
    headers = [cell.value for cell in powerbi_sheet[1]]
    canada_sheet.append(headers)

    # Determine the index for 'Project: Services Region'
    services_region_col_idx = headers.index("Project: Services Region") + 1

    # Initialize a list to keep track of rows to be removed
    rows_to_remove = []

    # Iterate through rows, find those with "Canada" in the specified column
    for row in powerbi_sheet.iter_rows(min_row=2):
        if "Canada" in str(row[services_region_col_idx - 1].value):
            rows_to_remove.append(row[0].row)
            canada_sheet.append([cell.value for cell in row])  # Append row to the Canada sheet

    # Remove rows containing "Canada" from the bottom up to avoid shifting indices
    for row_idx in reversed(rows_to_remove):
        powerbi_sheet.delete_rows(row_idx)

    # Save the modified workbook
    workbook.save("Updated_Partial_Revenue.xlsx")

if __name__ == "__main__":
    filter_canada_rows()




import openpyxl
from openpyxl import Workbook

def create_excel_with_je_sheet(filename, lookup_filename):
    # Load the lookup workbook and prepare the lookup dictionaries
    lookup_workbook = openpyxl.load_workbook(lookup_filename)
    lookup_sheet = lookup_workbook.active
    description_lookup = {}
    type_lookup = {}

    # Populate the lookup dictionaries from the lookup sheet
    for row in lookup_sheet.iter_rows(min_row=1, values_only=True):
        account_number = row[0]
        type_value = row[1]
        description_value = row[2]
        description_lookup[account_number] = description_value
        type_lookup[account_number] = type_value

    # Try to open an existing workbook, otherwise create a new one
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()

    # Define headers
    headers = ["Account Number", "Amount", "Description", "Reference 2", "Type"]

    # Check if the 'JE' sheet exists and clear it
    if 'JE' in workbook.sheetnames:
        je_sheet = workbook['JE']
        je_sheet.delete_rows(1, je_sheet.max_row)  # Clear all existing content
    else:
        je_sheet = workbook.create_sheet('JE')

    # List of Account Numbers
    account_numbers = [
        "002840.4110.10", "002844.4110.10", "002842.4110.10",
        "002842.4110.10", "002843.4110.10", "002841.4110.10"
    ]

    # Populate the account numbers and perform xlookup for description and type
    for account in account_numbers:
        description = description_lookup.get(account, "")
        account_type = type_lookup.get(account, "")
        je_sheet.append([account, "", description, "", account_type])

    # Insert headers at the top of the sheet by shifting all data down
    je_sheet.insert_rows(1)
    for col, header in enumerate(headers, start=1):
        je_sheet.cell(row=1, column=col).value = header

    # Save the workbook
    workbook.save(filename)

# Function call (uncomment in actual use)
create_excel_with_je_sheet("Updated_Partial_Revenue.xlsx", "lookup_gl.xlsx")





def update_cell(filename):
    workbook = openpyxl.load_workbook(filename)
    
    if 'JE' in workbook.sheetnames:
        je_sheet = workbook['JE']
        je_sheet['E4'] = "Professional Services"
        workbook.save(filename)
        
update_cell("Updated_Partial_Revenue.xlsx")

import openpyxl

def add_negative_sumifs_formula(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Ensure both "JE" and "PowerBi" sheets exist
    if "JE" not in workbook.sheetnames or "PowerBi" not in workbook.sheetnames:
        print("One or both required sheets are missing.")
        return

    # Access the sheets
    je_sheet = workbook["JE"]
    powerbi_sheet = workbook["PowerBi"]

    # Add the SUMIFS formula to column B in "JE", starting from row 2 to skip headers
    for row in range(2, je_sheet.max_row + 1):
        criteria = f'E{row}'  # Column E in "JE" as the criteria
        # SUMIFS formula in Excel, data from "PowerBi", make the result negative
        sumifs_formula = (
            f"=-SUMIFS(PowerBi!T:T, "  # Sum range in "PowerBi" sheet column T
            f"PowerBi!P:P, "  # Criteria range in "PowerBi" sheet column P
            f"JE!{criteria})"  # Criteria from column E in "JE"
        )
        je_sheet.cell(row, 2).value = sumifs_formula  # Set formula in column B of "JE"

    # Save the workbook
    workbook.save(file_path)
    print(f"SUMIFS formulas added and workbook saved as {file_path}")

# File path to the workbook
file_path = "Updated_Partial_Revenue.xlsx"

# Function call
add_negative_sumifs_formula(file_path)


#stop

import openpyxl

def process_excel_file(file_path):
    # Load the workbook and sheets
    wb = openpyxl.load_workbook(file_path)
    sheet_je = wb["JE"]
    sheet_powerbi = wb["PowerBi"]
    
    # Extract unique values from "PowerBi" column C
    unique_values = set()
    for cell in sheet_powerbi['C'][1:]:  # Skip the header row
        unique_values.add(cell.value)

    # Paste unique values into "JE" column D starting at row 8
    d_column_start = 8
    for i, value in enumerate(unique_values, start=d_column_start):
        sheet_je[f'D{i}'].value = value

    # Find the first blank row in column B for inserting SUMIFS formula
    b_column_first_blank = d_column_start + len(unique_values)

    # Insert SUMIFS formulas in column B in the "JE" tab
    for index in range(d_column_start, b_column_first_blank):
        criteria_address = f'D{index}'
        formula = f'=SUMIFS(PowerBi!$T:$T, PowerBi!$C:$C, JE!{criteria_address})'
        sheet_je[f'B{index}'].value = formula

    # Paste "2.2460.40" in column A from row 8 until it reaches a blank row in column D
    for index in range(d_column_start, b_column_first_blank):
        sheet_je[f'A{index}'].value = "1.1111.60"

    # Insert "NC 100%" in column C starting from row 8 until the value in column A is not "2.2460.40"
    index = d_column_start
    while sheet_je[f'A{index}'].value == "1.1111.60":
        sheet_je[f'C{index}'].value = "NC Partial Revenue"
        index += 1
    
    # Save the modified workbook
    wb.save('Updated_Partial_Revenue.xlsx')

# Uncomment the below line to execute the function with the specified file path
process_excel_file("Updated_Partial_Revenue.xlsx")


import openpyxl

def process_excel_file(file_path):
    # Load the workbook and sheets
    wb = openpyxl.load_workbook(file_path)
    sheet_je = wb["JE"]

    # Check if "002840.5110.10" already exists in column A
    already_exists = False
    for cell in sheet_je['A']:
        if cell.value is not None and cell.value.strip() == "002840.5110.10":
            already_exists = True
            print("Found existing '002840.5110.10'. No new row added.")
            break

    if not already_exists:
        print("Did not find '002840.5110.10'. Adding new row.")
        # Find next blank row in "JE" sheet
        next_blank_row = 1
        while sheet_je[f'A{next_blank_row}'].value is not None:
            next_blank_row += 1

        # Paste "002840.5110.10" in column A and "ASC 606 RecCost Hardware" in column C
        sheet_je[f'A{next_blank_row}'].value = "002840.5110.10"
        sheet_je[f'C{next_blank_row}'].value = "ASC 606 RecCost Hardware"

        
    # Save the modified workbook
    wb.save(file_path)

# Uncomment the below line to execute the function with the specified file path
process_excel_file("Updated_Partial_Revenue.xlsx")





import openpyxl

def update_je_sheet(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Ensure the "JE" sheet exists
    if "JE" not in workbook.sheetnames:
        print("The 'JE' sheet does not exist in the workbook.")
        return

    # Access the "JE" sheet
    je_sheet = workbook["JE"]

    # Search for the row containing the value "002840.5110.10" in column A
    target_value = "002840.5110.10"
    target_row = None
    for row_index, row in enumerate(je_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        if row[0] == target_value:
            target_row = row_index
            break

    # If found, apply the formula "=-B{row - 1} * 0.55" in column B of that row
    if target_row:
        je_sheet[f'B{target_row}'].value = f"=-B2 * 0.41"

    # Find the next blank row
    next_blank_row = 1
    while je_sheet[f'A{next_blank_row}'].value is not None:
        next_blank_row += 1

    # Insert "2.1399.40" to column A, negative value of the cell immediately above it in column B
    je_sheet[f'A{next_blank_row}'].value = "2.1399.40"
    je_sheet[f'B{next_blank_row}'].value = f"=-B{next_blank_row - 1}"
    je_sheet[f'C{next_blank_row}'].value = "Rec. Cost from RAR"

    # Save the workbook
    workbook.save(file_path)
    print(f"Workbook updated and saved as {file_path}")

# File path to the workbook
file_path = "Updated_Partial_Revenue.xlsx"

# Function call
update_je_sheet(file_path)



import openpyxl

def create_je_canada_sheet(main_file_path, lookup_file_path):
    # Load the main workbook and lookup workbook
    wb = openpyxl.load_workbook(main_file_path)
    lookup_wb = openpyxl.load_workbook(lookup_file_path)
    lookup_sheet = lookup_wb["US"]  # Adjusted for the "US" sheet
    
    # Create new sheet 'JE Canada' or clear it if it already exists
    if "JE Canada" in wb.sheetnames:
        je_canada = wb["JE Canada"]
        je_canada.delete_rows(1, je_canada.max_row)
    else:
        je_canada = wb.create_sheet("JE Canada")
    
    # Set headers
    headers = ["Account Number", "Amount", "Description", "Reference 2", "Type"]
    je_canada.append(headers)
    
    # Populate the Account Number column with predefined account numbers
    account_numbers = ["003840.4110.30", "003844.4110.30", "003842.4110.30", "003843.4110.30", "003841.4110.30"]
    for account in account_numbers:
        je_canada.append([account])

    # Build lookup dictionary from the lookup sheet
    lookup_dict_type = {}
    lookup_dict_desc = {}
    for row in lookup_sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers are in the first row
        lookup_dict_type[row[0]] = row[1]
        lookup_dict_desc[row[0]] = row[2]

    # Use the dictionary to assign values to 'Type' and 'Description' in 'JE Canada'
    for idx, account in enumerate(account_numbers, start=2):
        je_canada.cell(row=idx, column=5).value = lookup_dict_type.get(account, "NA")  # Column E for 'Type'
        je_canada.cell(row=idx, column=3).value = lookup_dict_desc.get(account, "NA")  # Column C for 'Description'
    
    # Save the workbook
    wb.save(main_file_path)

# Usage example
create_je_canada_sheet("Updated_Partial_Revenue.xlsx", "lookup_gl.xlsx")



import openpyxl

def update_formulas_in_je_canada(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Ensure the "JE Canada" sheet exists
    if "JE Canada" not in workbook.sheetnames:
        print("The 'JE Canada' sheet does not exist in the workbook.")
        return

    # Access the "JE Canada" sheet
    je_canada_sheet = workbook["PowerBi Canada"]

    # Iterate through each row in column T (20th column) starting from the second row to skip the header
    for row in range(2, je_canada_sheet.max_row + 1):
        # Example of updating a formula to reference cells in the same row, adjust according to actual needs
        # Let's say the formula multiplies values from column B and C in the same row
        je_canada_sheet.cell(row=row, column=20).value = f"=R{row} * U{row}"

    # Save the workbook
    workbook.save(file_path)
    print(f"Formulas in column T of 'JE Canada' have been updated. Workbook saved as {file_path}")

# File path to the workbook
file_path = "Updated_Partial_Revenue.xlsx"

# Function call
update_formulas_in_je_canada(file_path)


import openpyxl

def add_sumifs_to_je_canada(workbook_path):
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Ensure "JE Canada" sheet exists or create it
    if "JE Canada" not in wb.sheetnames:
        je_canada = wb.create_sheet("JE Canada")
        # Assuming we need to set headers or initialize something else if the sheet is newly created
        je_canada.append(["Account Number", "Amount", "Description", "Reference 2", "Type"])
    else:
        je_canada = wb["JE Canada"]

    # Ensure "PowerBi Canada" sheet exists
    if "PowerBi Canada" not in wb.sheetnames:
        print("Error: 'PowerBi Canada' sheet does not exist in the workbook.")
        return

    # Assuming account numbers or related data are already filled in JE Canada and we start formulas from row 2
    # Check and add formulas only if there's data in column E to use as criteria
    max_row = je_canada.max_row
    for row in range(2, max_row + 1):
        criteria_cell = je_canada.cell(row=row, column=5)  # Column E for criteria
        if criteria_cell.value is not None:
            sumifs_formula = f"=-SUMIFS('PowerBi Canada'!$T:$T, 'PowerBi Canada'!$P:$P, {criteria_cell.coordinate})"
            je_canada.cell(row=row, column=2).value = sumifs_formula  # Column B for the SUMIFS formula

    # Save changes to the workbook
    wb.save(workbook_path)

# Usage example
add_sumifs_to_je_canada("Updated_Partial_Revenue.xlsx")


import openpyxl

def update_je_canada_with_data(workbook_path):
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Access or create the necessary sheets
    if "JE Canada" not in wb.sheetnames:
        je_canada = wb.create_sheet("JE Canada")
    else:
        je_canada = wb["JE Canada"]

    if "PowerBi Canada" not in wb.sheetnames:
        print("Error: 'PowerBi Canada' sheet does not exist in the workbook.")
        return
    else:
        powerbi_canada = wb["PowerBi Canada"]

    # Extract unique values from column C of 'PowerBi Canada'
    unique_values = set()
    for cell in powerbi_canada['C']:
        if cell.value is not None and cell.row != 1:  # Exclude header and None values
            unique_values.add(cell.value)

    # Paste unique values into column D of 'JE Canada' starting from row 7
    start_row = 7
    for i, value in enumerate(unique_values, start=start_row):
        je_canada.cell(row=i, column=4).value = value  # Column 4 is column D

    # Add SUMIFS formula in column B of 'JE Canada'
    # Assuming formulas should be applied from row 7 onwards
    for i in range(start_row, start_row + len(unique_values)):
        criteria_cell = je_canada.cell(row=i, column=4)  # Criteria in column D
        sumifs_formula = f"=SUMIFS('PowerBi Canada'!$T:$T, 'PowerBi Canada'!$C:$C, {criteria_cell.coordinate})"
        je_canada.cell(row=i, column=2).value = sumifs_formula  # Column B for the SUMIFS formula

    # Save changes to the workbook
    wb.save(workbook_path)

# Usage example
update_je_canada_with_data("Updated_Partial_Revenue.xlsx")


import openpyxl

def update_sheet_based_on_condition(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
   
    # Check if the 'JE Canada' sheet exists
    if 'JE Canada' in workbook.sheetnames:
        sheet = workbook['JE Canada']
    else:
        print("Sheet 'JE Canada' does not exist.")
        return
   
    # Iterate through each row in the sheet starting from the first row with potential data
    for row in range(2, sheet.max_row + 1):
        # Check if there's a value in column D (column index 4)
        if sheet.cell(row=row, column=4).value is not None:
            # Insert "3.2460.40" in column A and "NC 100% Revenue" in column C
            sheet.cell(row=row, column=1).value = "3.2460.40"
            sheet.cell(row=row, column=3).value = "NC Partial Revenue"
   
    # Save the workbook with changes
    workbook.save(filename)
    print("Workbook updated successfully based on conditions in column D.")

# Function call (uncomment in actual use)
update_sheet_based_on_condition("Updated_Partial_Revenue.xlsx")


import openpyxl

def update_sheet_with_data(filename):
    # Load the workbook, assuming it exists.
    workbook = openpyxl.load_workbook(filename)

    # Access the 'JE Canada' sheet
    if 'JE Canada' in workbook.sheetnames:
        sheet = workbook['JE Canada']
    else:
        print("Sheet 'JE Canada' does not exist. Creating a new sheet.")
        sheet = workbook.create_sheet('JE Canada')

    # Find the next blank row in column A
    row_index = 1
    while sheet[f'A{row_index}'].value is not None:
        row_index += 1

    # Set the values in the next blank row
    sheet[f'A{row_index}'].value = "003840.5110.30"
    sheet[f'C{row_index}'].value = "ASC606 Rec Cost Hardware"
    # Insert the formula that calculates -B2 * 0.55 in column B
    sheet[f'B{row_index}'].value = "=-B2 * 0.55"

    # Save the workbook
    workbook.save(filename)
    print(f"Data added successfully in row {row_index} of 'JE Canada'.")

# Function call (uncomment in actual use)
update_sheet_with_data("Updated_Partial_Revenue.xlsx")



import openpyxl

def update_je_canada(filename):
    # Load the workbook, assuming it exists.
    workbook = openpyxl.load_workbook(filename)

    # Access or create the 'JE Canada' sheet
    if 'JE Canada' in workbook.sheetnames:
        sheet = workbook['JE Canada']
    else:
        print("Sheet 'JE Canada' does not exist. Creating a new sheet.")
        sheet = workbook.create_sheet('JE Canada')

    # Find the next blank row in column A
    row_index = 1
    while sheet[f'A{row_index}'].value is not None:
        row_index += 1

    # Paste the value "3.1399.40" in column A
    sheet[f'A{row_index}'].value = "3.1399.40"

    # Insert a formula in column B that equals the negative of the cell immediately above it
    if row_index > 1:  # Ensure there is a row above to reference
        sheet[f'B{row_index}'].value = f"=-B{row_index - 1}"
    else:
        sheet[f'B{row_index}'].value = "0"  # If it's the first row, no above cell to reference

    # Paste "Rec. Cost from RAR" in column C
    sheet[f'C{row_index}'].value = "Rec. Cost from RAR"

    # Save the workbook
    workbook.save(filename)
    print(f"Data added successfully in row {row_index} of 'JE Canada'.")

# Function call (uncomment in actual use)
update_je_canada("Updated_Partial_Revenue.xlsx")


import openpyxl

# Load the workbook and select the sheet
workbook = openpyxl.load_workbook('Updated_Partial_Revenue.xlsx')
sheet = workbook['PowerBi']

# Find the column index for "S/B Revenue", "R", and "U"
columns = {cell.value: cell.column for cell in sheet[1]}
sb_revenue_col = columns['S/B Revenue']
r_col = columns['Revenue (Gross)']
u_col = columns['ML Percentage Recognized']

# Iterate through each row starting from the second row (assuming the first row is the header)
for row in range(2, sheet.max_row + 1):
    r_value = sheet.cell(row=row, column=r_col).coordinate
    u_value = sheet.cell(row=row, column=u_col).coordinate
    formula = f'={r_value}*{u_value}'
    sheet.cell(row=row, column=sb_revenue_col).value = formula

# Save the workbook
workbook.save('Updated_Partial_Revenue.xlsx')


# New Addition

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated_Partial_Revenue.xlsx')

# Get the relevant sheets
variance_sheet = workbook['Variance']
powerbi_sheet = workbook['PowerBi Canada']

# Create a new sheet for "Variance Canada" if it doesn't already exist
if 'Variance Canada' not in workbook.sheetnames:
    variance_canada_sheet = workbook.create_sheet('Variance Canada')
else:
    variance_canada_sheet = workbook['Variance Canada']

# Copy headers from Variance sheet to Variance Canada sheet
for col_num, cell in enumerate(variance_sheet[1], 1):
    variance_canada_sheet.cell(row=1, column=col_num, value=cell.value)

# Identify the column indices for "SO#" in Variance and "JDE SO-JDE SO-" in PowerBi Canada
variance_headers = [cell.value for cell in variance_sheet[1]]
powerbi_headers = [cell.value for cell in powerbi_sheet[1]]

so_col_index = variance_headers.index("SO#") + 1
jde_so_col_index = powerbi_headers.index("JDE SO") + 1

# Function to find matching rows and perform copying and deletion
def match_and_transfer_rows():
    variance_rows_to_delete = []
    for var_row in variance_sheet.iter_rows(min_row=2, values_only=True):
        so_value = str(var_row[so_col_index - 1]).strip()
        for pbi_row in powerbi_sheet.iter_rows(min_row=2, values_only=True):
            if so_value == str(pbi_row[jde_so_col_index - 1]).strip():
                variance_canada_sheet.append(var_row)
                variance_rows_to_delete.append(var_row)
                break

    # Delete matching rows from Variance sheet
    for row in variance_rows_to_delete:
        for cell in variance_sheet.iter_rows(min_row=2):
            if str(cell[so_col_index - 1].value).strip() == str(row[so_col_index - 1]).strip():
                variance_sheet.delete_rows(cell[0].row)
                break

# Call the function to match and transfer rows
match_and_transfer_rows()

# Save the updated workbook
workbook.save('Updated_Partial_Revenue.xlsx')


# New Addition to add Canada variance tab.
import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated_Partial_Revenue.xlsx')

# Get the relevant sheets
variance_canada_sheet = workbook['Variance Canada']
powerbi_sheet = workbook['PowerBi Canada']
ml_sheet = workbook['Mavenlink']

# Find the column indices
variance_canada_headers = [cell.value for cell in variance_canada_sheet[1]]
so_col_index = variance_canada_headers.index("SO#") + 1

# Define column headers and insert them if they don't exist
if 'PowerBi' not in variance_canada_headers:
    variance_canada_headers.append('PowerBi')
    variance_canada_sheet.cell(row=1, column=len(variance_canada_headers), value='PowerBi')

if 'Mavenlink' not in variance_canada_headers:
    variance_canada_headers.append('Mavenlink')
    variance_canada_sheet.cell(row=1, column=len(variance_canada_headers), value='Mavenlink')

if 'Variance' not in variance_canada_headers:
    variance_canada_headers.append('Variance')
    variance_canada_sheet.cell(row=1, column=len(variance_canada_headers), value='Variance')

# Get column indices after adding new headers
powerbi_col_index = variance_canada_headers.index('PowerBi') + 1
mavenlink_col_index = variance_canada_headers.index('Mavenlink') + 1
variance_col_index = variance_canada_headers.index('Variance') + 1

# Iterate over the rows in Variance Canada sheet starting from the second row
for row in range(2, variance_canada_sheet.max_row + 1):
    so_value = variance_canada_sheet.cell(row=row, column=so_col_index).value

    # SUMIFS formula for the PowerBi column
    sumifs_formula = f'=SUMIFS(\'PowerBi Canada\'!T:T, \'PowerBi Canada\'!C:C, A{row})'
    variance_canada_sheet.cell(row=row, column=powerbi_col_index, value=sumifs_formula)

    # SUMIFS formula for the Mavenlink column
    xlookup_formula = f'=SUMIFS(\'Mavenlink\'!H:H, \'Mavenlink\'!D:D, A{row})'
    variance_canada_sheet.cell(row=row, column=mavenlink_col_index, value=xlookup_formula)

    # Variance formula
    variance_formula = f'={openpyxl.utils.get_column_letter(mavenlink_col_index)}{row} - {openpyxl.utils.get_column_letter(powerbi_col_index)}{row}'
    variance_canada_sheet.cell(row=row, column=variance_col_index, value=variance_formula)

# Save the updated workbook
workbook.save('Updated_Partial_Revenue.xlsx')

# Additional layer to ensure formulas are calculating correctly.

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated_Partial_Revenue.xlsx')

# Get the relevant sheets
variance_sheet = workbook['Variance']
powerbi_sheet = workbook['PowerBi']
ml_sheet = workbook['Mavenlink']

# Identify the column indices for "PowerBi" and "Mavenlink" in Variance sheet
variance_headers = [cell.value for cell in variance_sheet[1]]

if 'PowerBi' not in variance_headers:
    variance_headers.append('PowerBi')
    variance_sheet.cell(row=1, column=len(variance_headers), value='PowerBi')

if 'Mavenlink' not in variance_headers:
    variance_headers.append('Mavenlink')
    variance_sheet.cell(row=1, column=len(variance_headers), value='Mavenlink')

# Get column indices after adding new headers
powerbi_col_index = variance_headers.index('PowerBi') + 1
mavenlink_col_index = variance_headers.index('Mavenlink') + 1
criteria_col_index = variance_headers.index("SO#") + 1

# Iterate over the rows in Variance sheet starting from the second row
for row in range(2, variance_sheet.max_row + 1):
    # SUMIFS formula for the PowerBi column
    sumifs_powerbi_formula = f'=SUMIFS(\'PowerBi\'!T:T, \'PowerBi\'!C:C, A{row})'
    variance_sheet.cell(row=row, column=powerbi_col_index, value=sumifs_powerbi_formula)

    # SUMIFS formula for the Mavenlink column
    sumifs_mavenlink_formula = f'=SUMIFS(\'Mavenlink\'!H:H, \'Mavenlink\'!D:D, A{row})'
    variance_sheet.cell(row=row, column=mavenlink_col_index, value=sumifs_mavenlink_formula)

# Save the updated workbook
workbook.save('Updated_Partial_Revenue.xlsx')


# Additional layer to ensure variance column is calculating correctly.

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Updated_Partial_Revenue.xlsx')

# Get the relevant sheet
variance_sheet = workbook['Variance']

# Identify the column index for "Variance" and ensure it exists
variance_headers = [cell.value for cell in variance_sheet[1]]

if 'Variance' not in variance_headers:
    variance_headers.append('Variance')
    variance_sheet.cell(row=1, column=len(variance_headers), value='Variance')

# Get column indices after adding the new header
variance_col_index = variance_headers.index('Variance') + 1

# Iterate over the rows in Variance sheet starting from the second row
for row in range(2, variance_sheet.max_row + 1):
    # Formula to subtract the value in column B from the value in column C
    formula = f'=C{row} - B{row}'
    variance_sheet.cell(row=row, column=variance_col_index, value=formula)

# Save the updated workbook
workbook.save('Updated_Partial_Revenue.xlsx')


